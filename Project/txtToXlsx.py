import openpyxl

# Specify the input and output file paths
input_file_path = "human_chat.txt"
output_file_path = "output.xlsx"

# Open the input file with specified encoding
with open(input_file_path, "r", encoding="utf-8") as input_file:
    # Read lines from the file
    lines = input_file.readlines()

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Select the active sheet
sheet = workbook.active

# Write header
sheet["A1"] = "Question"
sheet["B1"] = "Answer"

# Iterate over lines and extract questions and answers
for row, i in enumerate(range(0, len(lines), 2), start=2):
    # Check if there are enough lines
    if i + 1 < len(lines):
        print(i)
        question = lines[i].split(": ")[1].strip()
        answer = lines[i + 1].split(": ")[1].strip()
        sheet.cell(row=row, column=1, value=question)
        sheet.cell(row=row, column=2, value=answer)
    else:
        print(f"Skipping line {i} as there is no corresponding answer.")

# Save the workbook
workbook.save(output_file_path)

print(f"Excel file '{output_file_path}' has been created.")
